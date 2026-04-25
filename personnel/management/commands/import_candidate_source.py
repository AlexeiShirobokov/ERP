from datetime import date, datetime
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from personnel.models import CandidateSourceRecord


def clean_text(value):
    if value is None:
        return ''

    if isinstance(value, datetime):
        return value.strftime('%d.%m.%Y')

    if isinstance(value, date):
        return value.strftime('%d.%m.%Y')

    if isinstance(value, float) and value.is_integer():
        if value > 20000:
            parsed = parse_excel_date(value)
            if parsed:
                return parsed.strftime('%d.%m.%Y')

        return str(int(value))

    if isinstance(value, int):
        if value > 20000:
            parsed = parse_excel_date(value)
            if parsed:
                return parsed.strftime('%d.%m.%Y')

        return str(value)

    return str(value).strip()


def normalize_header(value):
    """
    Нормализуем заголовок Excel, чтобы импорт не зависел от:
    - лишних пробелов;
    - точек;
    - регистра;
    - небольшой опечатки в слове "удостоверения".
    """
    text = clean_text(value).lower()
    text = text.replace('ё', 'е')
    text = text.replace('.', '')
    text = text.replace(',', '')
    text = text.replace('(', '')
    text = text.replace(')', '')
    text = ' '.join(text.split())
    return text


HEADER_ALIASES = {
    'source_number': [
        '№ п/п',
        '№',
        'номер',
    ],
    'source_date': [
        'дата',
    ],
    'full_name': [
        'фио',
        'фи о',
        'ф и о',
        'ф и о',
        'ф.и.о',
        'ф.и.о.',
    ],
    'birth': [
        'год рождения',
        'дата рождения',
        'год рождения дата рождения',
    ],
    'vacancy': [
        'вакансия',
    ],
    'phone': [
        'телефон',
        'контакты',
    ],

    # В Excel колонка называется с опечаткой:
    # "Квалификация, наличие удосоверения на сайте".
    # В карточке кандидата поле называется:
    # "Квалификация, наличие удостоверения на сайте".
    'qualification': [
        'квалификация наличие удостоверения на сайте',
        'квалификация наличие удосоверения на сайте',
        'квалификация',
    ],

    # Excel "Примечание" -> карточка "Примечание".
    'note': [
        'примечание',
    ],
    'medical_direction': [
        'направление на мо',
    ],

    # Excel "Примечание или причина отказа" ->
    # карточка "Примечание или причина отказа".
    'refusal_reason': [
        'примечание или причина отказа',
        'причина отказа',
    ],
    'accepted_date': [
        'принят дата',
        'принят',
        'дата принятия',
    ],
}


def build_column_map(header_row):
    """
    Возвращает словарь:
    {
        'qualification': 6,
        'note': 7,
        'refusal_reason': 9,
        ...
    }

    Индексы 0-based.
    """
    normalized_headers = [
        normalize_header(value)
        for value in header_row
    ]

    column_map = {}

    for field_name, aliases in HEADER_ALIASES.items():
        normalized_aliases = [
            normalize_header(alias)
            for alias in aliases
        ]

        for index, header in enumerate(normalized_headers):
            if not header:
                continue

            if header in normalized_aliases:
                column_map[field_name] = index
                break

    return column_map


def get_cell(row, column_map, field_name):
    index = column_map.get(field_name)

    if index is None:
        return None

    if index >= len(row):
        return None

    return row[index]


def parse_int(value):
    if value is None or value == '':
        return None

    if isinstance(value, int):
        return value

    if isinstance(value, float):
        return int(value)

    text = str(value).strip().replace(',', '.')

    if not text:
        return None

    try:
        return int(float(text))
    except ValueError:
        return None


def parse_excel_date(value):
    if value is None or value == '':
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date()
        except Exception:
            return None

    text = str(value).strip()

    if not text:
        return None

    for fmt in ['%d.%m.%Y', '%d.%m.%y', '%Y-%m-%d', '%d/%m/%Y', '%d/%m/%y']:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    return None


def parse_birth(value):
    if value is None or value == '':
        return None, None

    if isinstance(value, datetime):
        birth_date = value.date()
        return birth_date, birth_date.year

    if isinstance(value, date):
        return value, value.year

    if isinstance(value, (int, float)):
        if 1900 <= int(value) <= 2100:
            return None, int(value)

        birth_date = parse_excel_date(value)

        if birth_date:
            return birth_date, birth_date.year

    text = str(value).strip()

    if not text:
        return None, None

    if text.isdigit() and 1900 <= int(text) <= 2100:
        return None, int(text)

    birth_date = parse_excel_date(text)

    if birth_date:
        return birth_date, birth_date.year

    return None, None


class Command(BaseCommand):
    help = 'Импортирует Лист1 из Excel-файла вакансий в таблицу CandidateSourceRecord.'

    def add_arguments(self, parser):
        parser.add_argument(
            'xlsx_path',
            type=str,
            help='Путь к Excel-файлу, например "media/import/Вакансии 2026 (6).xlsx"',
        )

        parser.add_argument(
            '--sheet',
            type=str,
            default='Лист1',
            help='Название листа. По умолчанию: Лист1',
        )

        parser.add_argument(
            '--append',
            action='store_true',
            help='Не очищать старые записи перед импортом.',
        )

    def handle(self, *args, **options):
        xlsx_path = Path(options['xlsx_path'])
        sheet_name = options['sheet']

        if not xlsx_path.exists():
            raise CommandError(f'Файл не найден: {xlsx_path}')

        if not options['append']:
            deleted_count, _ = CandidateSourceRecord.objects.all().delete()
            self.stdout.write(
                self.style.WARNING(
                    f'Старые записи CandidateSourceRecord удалены: {deleted_count}'
                )
            )

        workbook = load_workbook(filename=xlsx_path, read_only=True, data_only=True)

        if sheet_name not in workbook.sheetnames:
            available = ', '.join(workbook.sheetnames)
            raise CommandError(
                f'Лист "{sheet_name}" не найден. Доступные листы: {available}'
            )

        sheet = workbook[sheet_name]

        rows = sheet.iter_rows(values_only=True)

        try:
            header_row = next(rows)
        except StopIteration:
            workbook.close()
            raise CommandError('Лист пустой.')

        column_map = build_column_map(header_row)

        required_fields = [
            'full_name',
            'qualification',
            'note',
            'refusal_reason',
        ]

        missing_fields = [
            field_name
            for field_name in required_fields
            if field_name not in column_map
        ]

        if missing_fields:
            workbook.close()
            visible_headers = ', '.join(clean_text(value) for value in header_row if clean_text(value))

            raise CommandError(
                'Не найдены обязательные колонки: '
                f'{", ".join(missing_fields)}. '
                f'Найденные заголовки: {visible_headers}'
            )

        self.stdout.write(
            self.style.SUCCESS(
                'Колонки найдены: '
                f'qualification={column_map.get("qualification") + 1}, '
                f'note={column_map.get("note") + 1}, '
                f'refusal_reason={column_map.get("refusal_reason") + 1}'
            )
        )

        created_count = 0
        skipped_count = 0

        for row_index, row in enumerate(rows, start=2):
            full_name = clean_text(get_cell(row, column_map, 'full_name'))

            if not full_name:
                skipped_count += 1
                continue

            birth_date, birth_year = parse_birth(
                get_cell(row, column_map, 'birth')
            )

            CandidateSourceRecord.objects.create(
                source_row=row_index,
                source_number=parse_int(get_cell(row, column_map, 'source_number')),
                source_date=parse_excel_date(get_cell(row, column_map, 'source_date')),
                full_name=full_name,
                birth_date=birth_date,
                birth_year=birth_year,
                vacancy=clean_text(get_cell(row, column_map, 'vacancy')),
                phone=clean_text(get_cell(row, column_map, 'phone')),

                # Маппинг, который нужен:
                # Excel "Квалификация, наличие удосоверения на сайте"
                # -> карточка "Квалификация, наличие удостоверения на сайте"
                qualification=clean_text(get_cell(row, column_map, 'qualification')),

                # Excel "Примечание" -> карточка "Примечание"
                note=clean_text(get_cell(row, column_map, 'note')),

                medical_direction=clean_text(get_cell(row, column_map, 'medical_direction')),

                # Excel "Примечание или причина отказа"
                # -> карточка "Примечание или причина отказа"
                refusal_reason=clean_text(get_cell(row, column_map, 'refusal_reason')),

                accepted_date=parse_excel_date(get_cell(row, column_map, 'accepted_date')),
                import_file_name=xlsx_path.name,
            )

            created_count += 1

        workbook.close()

        self.stdout.write(
            self.style.SUCCESS(
                f'Импорт завершён. Создано: {created_count}. Пропущено: {skipped_count}.'
            )
        )
