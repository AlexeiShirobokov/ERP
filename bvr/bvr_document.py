# -*- coding: utf-8 -*-
"""
Генератор проекта массового взрыва: Excel (.xlsx) и PDF.
Формирует полный проект — обложка, акт, распорядок, таблицы, расчёты,
приказ, распоряжение, расчёт безопасных расстояний.
"""
import os
import datetime
from .bvr_calc import MONTHS, depth_groups

# ---------- ФОРМАТИРОВАНИЕ ----------
def fmt(x, d=2):
    if x is None:
        return "0"
    s = f"{x:,.{d}f}".replace(",", " ").replace(".", ",")
    return s

def fmt0(x):
    return fmt(x, 0)

def date_parts(params):
    d = params.get("data") or datetime.date.today()
    return d.day, MONTHS[d.month - 1], d.year

def date_str(params):
    dd, mm, yy = date_parts(params)
    return f"« {dd} » {mm} {yy} г."


# ==================== EXCEL ====================
def _summary_values(calc, summary):
    """Сводные count/метраж/средняя глубина: из паспорта (summary) либо из грида."""
    s = summary or {}
    n_disp = int(s.get("wells") or calc["n"])
    drill_disp = float(s.get("meters") or calc["sumD"])
    avg_disp = float(s.get("avg") or calc["sr_glub"])
    return n_disp, drill_disp, avg_disp


def build_excel(calc, path, summary=None):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    p = calc["params"]
    n_disp, drill_disp, avg_disp = _summary_values(calc, summary)
    wb = openpyxl.Workbook()
    thin = Side(style="thin", color="B0B8C4")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="1D4E89")
    sec_fill = PatternFill("solid", fgColor="EEF3F9")
    hdr_font = Font(bold=True, color="FFFFFF")
    bold = Font(bold=True)

    # --- Лист «Параметры» ---
    ws = wb.active
    ws.title = "Параметры"
    rows = [
        ("ПРОЕКТ МАССОВОГО ВЗРЫВА — расчёт БВР", "", ""),
        ("", "", ""),
        ("Предприятие", "", p["pred"]),
        ("Месторождение", "", p["mest"]),
        ("Блок", "", p["blok"]),
        ("Дата взрыва", "", date_str(p)),
        ("", "", ""),
        ("Параметр", "Ед. изм.", "Значение"),
        ("Расположение скважин", "", p["raspol"]),
        ("Плотность ВВ в заряде", "г/см³", p["plotvv"]),
        ("Условный диаметр скважин", "м", p["diam"]),
        ("КИС", "", p["kis"]),
        ("Расход селитры на 1 п.м", "кг/пог.м", calc["raskh_sel"]),
        ("ЛСПП", "м", calc["lspp"]),
        ("Коэффициент сближения зарядов", "", calc["ksbl"]),
        ("Расстояние между скважинами", "м", p["a"]),
        ("Расстояние между рядами", "м", calc["b"]),
        ("Количество рядов", "шт", p["ryad"]),
        ("Количество скважин", "шт", n_disp),
        ("Площадь массива", "м²", p["ploshad"]),
        ("Объём взрываемого массива", "м³", round(calc["objem_massiva"], 1)),
        ("Общий объём бурения", "п.м", round(drill_disp, 1)),
        ("Средняя глубина скважин", "м", round(avg_disp, 2)),
        ("Средняя высота уступа", "м", round(calc["sr_hust"], 2)),
        ("Средняя длина заряда", "м", round(calc["sr_zar"], 2)),
        ("Удельный расход ВВ", "кг/м³", round(calc["ud_vv"], 3)),
        ("", "", ""),
        ("РАСЧЁТ ВВ", "Ед.", "Значение"),
        ("Аммиачная селитра", "кг", round(calc["mas_as"], 2)),
        ("Дизтопливо", "кг", round(calc["mas_dt_kg"], 2)),
        ("Дизтопливо", "л", round(calc["mas_dt_l"], 2)),
        (_boevik_name(p), "кг", round(calc["mas_pt"], 2)),
        ("Общая масса заряда", "кг", round(calc["obsh_massa"], 2)),
        ("Средняя масса заряда в скважине", "кг", round(calc["sr_massa"], 2)),
        ("", "", ""),
        ("СРЕДСТВА ИНИЦИИРОВАНИЯ", "Ед.", "Значение"),
        ("Искра-С (500 мс)", "шт", calc["iskra_s"]),
        ("Искра-П (67 мс)", "шт", calc["iskra67"]),
        ("Искра-П (42 мс)", "шт", calc["iskra42"]),
        ("Искра-В (Старт-600м)", "шт", calc["iskraV"]),
        ("", "", ""),
        ("БЕЗОПАСНЫЕ РАССТОЯНИЯ", "Ед.", "Значение"),
        ("Опасная зона для людей", "м", calc["zona_ludi"]),
        ("Опасная зона для оборудования", "м", calc["zona_obor"]),
    ]
    for r in rows:
        ws.append(r)
    ws["A1"].font = Font(bold=True, size=13, color="1D4E89")
    for cell in ("A8", "B8", "C8", "A28", "B28", "C28", "A37", "B37", "C37",
                 "A43", "B43", "C43"):
        ws[cell].font = hdr_font
        ws[cell].fill = hdr_fill
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 22

    # --- Лист «Каталог скважин» ---
    ws = wb.create_sheet("Каталог скважин")
    head = ["№", "Ряд-№ скв", "Глубина, м", "Высота уступа, м", "Перебур, м",
            "Высота колонки, м", "Забойка, м", "АС, кг", "ДТ, кг", "ДТ, л",
            "Гранулит, кг", "Vгм, м³"]
    ws.append(head)
    for c in ws[1]:
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for i, w in enumerate(calc["wells"], 1):
        ws.append([i, f"{w['ryad']}-{w['skv']}", round(w["d"], 2),
                   round(w["H"], 2), round(w["I"], 2), round(w["N"], 2),
                   round(w["zaboi"], 2), round(w["K"], 2), round(w["L"], 2),
                   round(w["M"], 2), round(w["E"], 2), round(w["J"], 2)])
    tot = ["", "ИТОГО", round(calc["sumD"], 2), round(calc["sumH"], 2),
           round(calc["sumI"], 2), round(calc["sumN"], 2),
           round(calc["n"] * p["zaboi"], 2), round(calc["mas_as"], 2),
           round(calc["mas_dt_kg"], 2), round(calc["mas_dt_l"], 2),
           round(calc["mas_as"] + calc["mas_dt_kg"], 2),
           round(calc["objem_massiva"], 2)]
    ws.append(tot)
    for c in ws[ws.max_row]:
        c.font = bold
        c.fill = PatternFill("solid", fgColor="FFF4E6")
    for col in "ABCDEFGHIJKL":
        ws.column_dimensions[col].width = 13

    # --- Лист «Зарядная карта» ---
    ws = wb.create_sheet("Зарядная карта")
    max_skv = max((w["skv"] for w in calc["wells"]), default=1)
    max_ryad = max((w["ryad"] for w in calc["wells"]), default=1)
    ws.append(["Ряд \\ Скв."] + list(range(1, max_skv + 1)))
    grid = {}
    for w in calc["wells"]:
        grid[(w["ryad"], w["skv"])] = w["d"]
    for r in range(1, max_ryad + 1):
        ws.append([r] + [grid.get((r, c), "") for c in range(1, max_skv + 1)])
    for c in ws[1]:
        c.font = hdr_font
        c.fill = hdr_fill

    # --- Лист «Таблица параметров» ---
    ws = wb.create_sheet("Таблица параметров")
    ws.append(["№", "Глубина, м", "Кол-во скв.", "Масса заряда, кг",
               "Длина заряда, м", "Длина забойки, м"])
    for c in ws[1]:
        c.font = hdr_font
        c.fill = hdr_fill
    for i, g in enumerate(depth_groups(calc), 1):
        mass = (g["sK"] + g["sL"]) / g["n"] + p["boevik"]
        ws.append([i, round(g["d"], 2), g["n"], round(mass, 2),
                   round(g["sN"] / g["n"], 2), round(g["sZ"] / g["n"], 2)])
    for col in "ABCDEF":
        ws.column_dimensions[col].width = 18

    wb.save(path)
    return path


# ==================== PDF ====================
def _register_font():
    """Подбирает TTF-шрифт с поддержкой кириллицы."""
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    candidates = [
        ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
         "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
        ("/usr/share/fonts/truetype/liberation/LiberationSerif-Regular.ttf",
         "/usr/share/fonts/truetype/liberation/LiberationSerif-Bold.ttf"),
        ("/Library/Fonts/Arial Unicode.ttf",
         "/Library/Fonts/Arial Unicode.ttf"),
    ]
    for reg, bold in candidates:
        if os.path.exists(reg):
            pdfmetrics.registerFont(TTFont("BVR", reg))
            pdfmetrics.registerFont(TTFont("BVR-B", bold if os.path.exists(bold) else reg))
            return "BVR", "BVR-B"
    return "Helvetica", "Helvetica-Bold"


def _boevik_name(p):
    """Метка боевика: масса 0.5 кг -> ПТ-П-500, 2.25 кг -> ПТ-П-2250."""
    try:
        return "ПТ-П-%d" % int(round(float(p.get("boevik", 0)) * 1000))
    except (TypeError, ValueError):
        return "ПТ-П"


def _participants(p):
    """Список участников листа «Ознакомление» -> [(ФИО, профессия), ...].

    Источник — строковое поле ``participants``: по одному участнику на строку,
    ФИО и профессия разделены «;» (либо табом/запятой).
    """
    out = []
    for line in (p.get("participants") or "").splitlines():
        line = line.strip()
        if not line:
            continue
        sep = ";" if ";" in line else ("\t" if "\t" in line else ",")
        fio, _, prof = line.partition(sep)
        out.append((fio.strip(), prof.strip()))
    return out


def build_pdf(calc, path, summary=None, passport_path=None):
    """Проект массового взрыва (PDF) — воспроизводит листы Excel-эталона 1:1.

    Листы: Обложка, Акт, Ознакомление, Распорядок, Таблица параметров,
    Инструктаж (2 листа), Приказ, Распоряжение о постах, Расчёт параметров
    (3 листа), Безопасные расстояния (3 листа).
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    PageBreak, Table, TableStyle, KeepTogether)

    FONT, FONT_B = _register_font()
    p = calc["params"]
    n_disp, drill_disp, avg_disp = _summary_values(calc, summary)
    dd, mon, yy = date_parts(p)
    D = '« %s » %s %s г.' % (dd, mon, yy)
    bname = _boevik_name(p)

    pred = p["pred"]; mest = p["mest"]; blok = p["blok"]; prisk = p["prisk"]
    gling = p["gling"]; vzr = p["vzryvnik"]; injot = p["injot"]; dispet = p["dispet"]
    # Ответственный руководитель ВР во всём проекте — начальник участка.
    nu = 'Начальник участка "%s" %s' % (mest, p["nachuch"])

    # ---- Стили ----
    body = ParagraphStyle("body", fontName=FONT, fontSize=10.5, leading=14.5,
                          alignment=4, spaceAfter=3)
    item = ParagraphStyle("item", parent=body, spaceAfter=4)
    center = ParagraphStyle("center", parent=body, alignment=1)
    cbold = ParagraphStyle("cbold", parent=center, fontName=FONT_B)
    bld = ParagraphStyle("bld", parent=body, fontName=FONT_B)
    right = ParagraphStyle("right", parent=body, alignment=2)
    h1 = ParagraphStyle("h1", fontName=FONT_B, fontSize=14, leading=18,
                        alignment=1, spaceAfter=6)
    cover_big = ParagraphStyle("cb", fontName=FONT_B, fontSize=26, leading=34,
                               alignment=1)
    small = ParagraphStyle("small", parent=body, fontSize=8.5, leading=11)
    csmall = ParagraphStyle("csmall", parent=small, alignment=1)

    story = []
    def P(t, s=body): story.append(Paragraph(t, s))
    def SP(h=6): story.append(Spacer(1, h))

    def underline_row(num, label, value, unit, lw=250):
        t = Table([["%s.   %s" % (num, label), value, unit]],
                  colWidths=[lw, 150, 70])
        t.setStyle(TableStyle([
            ("FONT", (0, 0), (-1, -1), FONT, 10.5),
            ("FONT", (1, 0), (1, 0), FONT_B, 10.5),
            ("ALIGN", (1, 0), (1, 0), "CENTER"),
            ("LINEBELOW", (1, 0), (1, 0), 0.7, colors.black),
            ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
        ]))
        return t

    # ============================ ОБЛОЖКА ============================
    SP(70)
    P("АКЦИОНЕРНОЕ ОБЩЕСТВО", cbold)
    P('"%s"' % pred.replace("АО", "").replace('"', "").strip().upper(), cbold)
    SP(120)
    P("П Р О Е К Т", cover_big)
    SP(6)
    P("МАССОВОГО ВЗРЫВА", ParagraphStyle("c2", parent=cover_big, fontSize=18))
    SP(46)
    P("месторождение: %s" % mest, center)
    P("Блок: %s" % blok, center)
    SP(150)
    P(p["np"], center)
    P("%s г." % yy, center)
    story.append(PageBreak())

    # ============================== АКТ ==============================
    P(pred, cbold)
    SP(8)
    P("А К Т", h1)
    P("о готовности блока к заряжанию", center)
    SP(10)
    hdr = Table([["Россыпное месторождение:", mest, "блок:", blok]],
                colWidths=[170, 130, 45, 150])
    hdr.setStyle(TableStyle([("FONT", (0, 0), (-1, -1), FONT, 10.5),
                             ("FONT", (1, 0), (1, 0), FONT_B, 10.5),
                             ("FONT", (3, 0), (3, 0), FONT_B, 10.5)]))
    story.append(hdr)
    SP(2)
    dr = Table([['"______"   %s   %s г.' % (mon, yy), prisk]],
               colWidths=[260, 235])
    dr.setStyle(TableStyle([("FONT", (0, 0), (-1, -1), FONT, 10.5),
                            ("ALIGN", (1, 0), (1, 0), "RIGHT")]))
    story.append(dr)
    SP(10)
    P("Мы, нижеподписавшиеся, ответственный руководитель ВР %s, маркшейдер "
      "________________, составили настоящий акт о том, что блок <b>%s</b> "
      "полностью забурен и подготовлен к заряжанию в следующих параметрах:"
      % (p["nachuch"], blok))
    SP(6)
    story.append(underline_row("1", "Средняя глубина скважин", fmt(avg_disp), "м"))
    story.append(underline_row("2", "Площадь блока", fmt(p["ploshad"] / 1000, 3), "тыс.м2"))
    story.append(underline_row("3", "Количество скважин", str(n_disp), "шт."))
    story.append(underline_row("4", "Объем торфов на взрыв", fmt(calc["objem_massiva"] / 1000, 2), "тыс.м3"))
    SP(8)
    P("Скважины пробурены в соответствии с проектом и очищены. Блок очищен "
      "от посторонних предметов и металлолома.")
    SP(26)
    sig = Table([["Ответственный:", ""],
                 ["руководитель ВР", p["nachuch"]],
                 ["Маркшейдер", ""]],
                colWidths=[250, 245])
    sig.setStyle(TableStyle([
        ("FONT", (0, 0), (-1, -1), FONT, 10.5),
        ("BOX", (0, 0), (-1, -1), 1.1, colors.HexColor("#2f7d4f")),
        ("LINEBELOW", (1, 1), (1, 1), 0.7, colors.black),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
    ]))
    story.append(sig)
    story.append(PageBreak())

    # ====================== ОЗНАКОМЛЕНИЕ (лист 2) ====================
    P("С проектом массового взрыва ознакомлены и проинструктированы:", bld)
    SP(6)
    parts = _participants(p)
    data = [["№ п/п", "Ф.И.О.", "Профессия", "Роспись"]]
    total = max(19, len(parts) + 2)
    for i in range(1, total + 1):
        fio, prof = parts[i - 1] if i <= len(parts) else ("", "")
        data.append([str(i), fio, prof, ""])
    t = Table(data, colWidths=[45, 220, 150, 100])
    t.setStyle(TableStyle([
        ("FONT", (0, 0), (-1, -1), FONT, 9.5),
        ("FONT", (0, 0), (-1, 0), FONT_B, 9.5),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EEF3F9")),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("TOPPADDING", (0, 1), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
    ]))
    story.append(t)
    story.append(PageBreak())

    # ========================== РАСПОРЯДОК ===========================
    P("«УТВЕРЖДАЮ»<br/>Главный инженер %s<br/>________________ %s" % (pred, gling), right)
    SP(6)
    P("Р А С П О Р Я Д О К", h1)
    P("проведения массового взрыва", center)
    SP(6)
    P("Россыпное месторождение: <b>%s</b>   блок: <b>%s</b>" % (mest, blok))
    items = [
        'Дата взрыва: %s' % D,
        'Место взрыва: месторождение %s, %s' % (mest, prisk),
        'Время взрыва: с %s до %s по местному времени' % (p["vz1"], p["vz2"]),
        'Зарядка производится: %s с %s до %s' % (D, p["zar1"], p["zar2"]),
        'Способы инициирования зарядов, взрывной сети: электрический',
        'Общее количество взрываемых скважин (шт.): <b>%d</b>' % n_disp,
        'Способ взрывания: инициирующий импульс от капсюля-детонатора '
        'Искра-В (Старт-600м) устройства Искра-С, с интервалом замедления: 500 мс',
        'Тип замедлителей: ИСКРА-П (42; 67) мс',
        'Схема взрывания с указанием величин интервалов замедлений: %s' % p["shema"],
        'Порядок монтажа взрывной сети: к источнику тока',
        'Место расположения взрывной станции: согласно схемы',
        'Опасная зона: для людей — %s м; для оборудования — %s м; '
        'для сооружений — %s м.' % (calc["zona_ludi"], calc["zona_obor"], p["zoso"]),
        'Объекты находящиеся в опасной зоне: %s' % p["obj"],
        'Мероприятия по предотвращению повреждений охраняемых объектов: %s' % p["meri"],
        'Оборудование отводится от ближайшей скважины: экскаваторы на %s м; '
        'буровые станки на %s м.' % (calc["zona_obor"], calc["zona_obor"]),
        'Схема расстановки постов охраны опасной зоны прилагается.',
        'Ответственным руководителем массового взрыва назначен: %s '
        '(должность, фамилия, инициалы).' % nu,
        'Подвозка взрывчатых материалов к месту взрыва производится '
        'специализированным автомобилем: %s, водитель: %s. Сопровождающее лицо: '
        '%s.' % (p["avtoVM"], p["voditel"] or "____________", nu),
        'Для очистки скважин перед заряжанием используется: СБШ-250. '
        'Обслуживаемая бригада: ____________.',
    ]
    for i, it in enumerate(items, 1):
        P("%d. %s" % (i, it), item)
    story.append(PageBreak())

    # ======================= ТАБЛИЦА ПАРАМЕТРОВ ======================
    P("Т А Б Л И Ц А", h1)
    P("параметров взрывных работ", center)
    P("на россыпном месторождении золота %s, блок %s на взрыв %s" % (mest, blok, D), csmall)
    SP(6)
    groups = depth_groups(calc)
    diam_mm = str(round(p["diam"] * 1000))
    head1 = ["№\nп/п", "Глубина\nскважин, м", "", "Количество\nскважин", "",
             "Расстоя-\nние м/у\nскваж., м", "", "Расстоя-\nние м/у\nрядами, м", "",
             "Диаметр\nскваж.,\nмм", "", "Масса\nзаряда в\nскваж., кг", "",
             "Длина\nзаряда,\nм", "", "Длина\nзабойки,\nм", "", "Приме-\nчание"]
    head2 = ["", "Р", "Ф", "Р", "Ф", "Р", "Ф", "Р", "Ф", "Р", "Ф",
             "Р", "Ф", "Р", "Ф", "Р", "Ф", ""]
    tdata = [head1, head2]
    for i, g in enumerate(groups, 1):
        mass = (g["sK"] + g["sL"]) / g["n"] + p["boevik"]
        tdata.append([str(i), fmt(g["d"], 1), "", str(g["n"]), "", fmt(p["a"], 1), "",
                      fmt(calc["b"], 1), "", diam_mm, "", fmt(mass, 1), "",
                      fmt(g["sN"] / g["n"], 1), "", fmt(g["sZ"] / g["n"], 1), "", ""])
    for _ in range(max(0, 11 - len(groups))):
        tdata.append([""] * 18)
    cw = [20, 30, 16, 34, 16, 30, 16, 30, 16, 30, 16, 40, 18, 30, 16, 30, 16, 40]
    t = Table(tdata, colWidths=cw, repeatRows=2)
    st = [
        ("FONT", (0, 0), (-1, -1), FONT, 6.5),
        ("FONT", (0, 0), (-1, 1), FONT_B, 6.5),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.black),
        ("BACKGROUND", (0, 0), (-1, 1), colors.HexColor("#EEF3F9")),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("SPAN", (0, 0), (0, 1)),    # № п/п
        ("SPAN", (17, 0), (17, 1)),  # Примечание
    ]
    for c in range(1, 17, 2):
        st.append(("SPAN", (c, 0), (c + 1, 0)))
    t.setStyle(TableStyle(st))
    story.append(t)
    SP(8)
    asn = calc["mas_as"] / calc["n"] if calc["n"] else 0
    dtn = calc["mas_dt_kg"] / calc["n"] if calc["n"] else 0
    blk1 = ("<b>Расчетная формула усредненного веса заряда:</b><br/>"
            "1. Боевик %s — %s кг<br/>2. Селитра / скв. — %s кг<br/>"
            "3. Дизельное топливо — %s кг<br/><b>Итого: %s кг</b>"
            % (bname, fmt(p["boevik"]), fmt(asn), fmt(dtn), fmt(calc["sr_massa"])))
    blk2 = ("<b>Общая масса заряда:</b><br/>"
            "1. Гранулит М (АС+ДТ) — %s кг<br/>2. %s — %s кг<br/>"
            "<b>Итого ВМ: %s кг</b>"
            % (fmt(calc["mas_as"] + calc["mas_dt_kg"], 1), bname,
               fmt(calc["mas_pt"], 1), fmt(calc["obsh_massa"], 1)))
    blk3 = ("<b>Потребляемое кол-во СИ:</b><br/>"
            "1. ИСКРА-В-Старт — %s<br/>2. ИСКРА-С (500мс) — %s<br/>"
            "3. ИСКРА-П (42 мс) — %s<br/>4. ИСКРА-П (67 мс) — %s"
            % (calc["iskraV"], calc["iskra_s"], calc["iskra42"], calc["iskra67"]))
    foot = Table([[Paragraph(blk1, small), Paragraph(blk2, small), Paragraph(blk3, small)]],
                 colWidths=[180, 165, 150])
    foot.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
    story.append(foot)
    SP(10)
    P("Расчет составил: %s &nbsp;&nbsp;&nbsp; %s" % (nu, D), small)
    P("Расчет проверил: Главный инженер %s &nbsp;&nbsp;&nbsp; %s" % (gling, D), small)
    story.append(PageBreak())

    # ===================== ИНСТРУКТАЖ (лист 5) =======================
    P("Инструктаж работников провёл: <b>%s</b>" % nu)
    P("(кем, когда)", csmall)
    SP(4)
    instr5 = [
        'Непосредственное руководство работой персонала при подготовке и '
        'проведении массового взрыва, учет и сохранность ВМ при заряжании '
        'скважин возложить на: %s.' % nu,
        'Ответственным за вывод людей с территории запретной и опасной зон '
        'назначен: %s.' % nu,
        'Ответственным за заряжание и монтаж взрывной (электровзрывной) сети '
        'назначен: Старший взрывник %s.' % vzr,
        'Ответственным за вывод внутрикарьерного транспорта из запретной и '
        'опасной зон назначен: %s.' % nu,
        'Ответственным за охрану запретной и опасной зон назначен: %s.' % nu,
        'Ответственным за отключение электроэнергии, удаление в безопасное '
        'место эл/оборудования перед взрывом, а также за проверку и подключение '
        'ее после взрыва назначен: %s.' % nu,
        'Ответственным за подачу звуковых и световых сигналов назначен: %s.' % nu,
        'Ответственным за оповещение соседних предприятий, подразделений '
        'назначен: %s.' % nu,
        'Подача сигналов проводится по распоряжению. Исполнитель: %s. Звуковая '
        'сирена, установленная на зарядно-смесительной машине МЗ-3Б.' % p["nachuch"],
        'После выставления постов подается предупредительный сигнал: один '
        'продолжительный звуковой, оператором МЗ-3Б.',
        'По указанию ответственного за вывод людей все люди, не занятые '
        'заряжанием, должны удалиться за пределы опасной зоны: место расположения '
        '— столовая вахтового посёлка, согласно схемы.',
        'Осуществляются перечисленные в распорядке проведения массового взрыва '
        'дополнительные меры безопасности, связанные с вводом запретной зоны: '
        'запрещается проход не задействованных на зарядке работников, запрещен '
        'проезд техники.',
        'Заряжание скважин осуществляет: Старший взрывник %s.' % vzr,
        'Место сбора лиц, выполняющих заряжание, перед выходом из запретной '
        'зоны: пост №1.',
        'По завершению заряжания выставляются посты охраны опасной зоны.',
    ]
    for i, it in enumerate(instr5, 1):
        P("%d. %s" % (i, it), item)
    story.append(PageBreak())

    # ===================== ИНСТРУКТАЖ (лист 7) =======================
    P("1. Укладку в заряды боевиков с капсюлем, монтаж электровзрывной сети "
      "осуществляет: Старший взрывник %s, под руководством ответственного "
      "руководителя ВР %s." % (vzr, nu), item)
    P("2. Боевой сигнал: два длинных звуковых, оператором МЗ-3Б.", item)
    P("3. После подачи боевого сигнала производится взрыв: неэлектрический "
      "(способ взрывания).", item)
    P("4. Сигнал отбой: три коротких после осмотра места взрыва, после "
      "получения указания от ответственного руководителя ВР %s." % nu, item)
    P("5. Время проветривания и допуска людей в карьер, к месту взрыва — "
      "00 час 30 мин.", item)
    P("6. С распорядком проведения массового взрыва ознакомлены:", bld)
    nuu = p["nachuch"]
    soglas = [
        ("Ответственный руководитель взрывных работ", nuu),
        ("Ответственный за вывод людей из опасной и запретной зон", nuu),
        ("Ответственный руководитель ВР в смене", nuu),
        ("Ответственный за заряжание и монтаж взрывной сети", vzr),
        ("Ответственный за вывод внутрикарьерного транспорта", nuu),
        ("Ответственный за отключение электроэнергии", nuu),
        ("Ответственный за охрану запретной и опасной зон", nuu),
        ("Ответственный за подачу сигналов", nuu),
        ("Ответственный за оповещение соседних предприятий", injot),
    ]
    sdata = [["%s:" % r, fio, "____________"] for r, fio in soglas]
    ts = Table(sdata, colWidths=[300, 110, 90])
    ts.setStyle(TableStyle([
        ("FONT", (0, 0), (-1, -1), FONT, 9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
    ]))
    story.append(ts)
    SP(10)
    P("Распорядок проведения массового взрыва составил: %s &nbsp;&nbsp; %s" % (nu, D), small)
    P("Распорядок проведения массового взрыва проверил: Главный инженер %s &nbsp;&nbsp; %s" % (gling, D), small)
    story.append(PageBreak())

    # ============================ ПРИКАЗ =============================
    # Компактные стили — приказ (17 пунктов) должен умещаться на одном листе.
    pst = ParagraphStyle("pst", parent=body, fontSize=9, leading=11, spaceAfter=2)
    pcenter = ParagraphStyle("pcenter", parent=pst, alignment=1)
    pcbold = ParagraphStyle("pcbold", parent=pcenter, fontName=FONT_B)
    ph1 = ParagraphStyle("ph1", fontName=FONT_B, fontSize=13, leading=15,
                         alignment=1, spaceAfter=4)
    prikaz = []
    def AP(t, s=pst): prikaz.append(Paragraph(t, s))
    AP("П Р И К А З   №   %s" % p["prikaz"], ph1)
    AP('от %s &nbsp;&nbsp;&nbsp; %s' % (D, p["np"]), pcenter)
    AP("О производстве массового взрыва %s" % D, pcbold)
    AP("на месторождении %s, блок № %s." % (mest, blok), pcenter)
    AP("Согласно графику производства массовых взрывов и проекта горных работ "
       "на участке россыпного золота месторождение %s, блок %s, %s, %s "
       "с %s до %s часов будет произведен массовый взрыв."
       % (mest, blok, prisk, D, p["vz1"], p["vz2"]))
    AP("В целях обеспечения безопасности работ и сохранности ВМ,")
    AP("ПРИКАЗЫВАЮ:", pcbold)
    pitems = [
        'Проект производства массового взрыва на %s, разработанный на основании '
        'типового проекта, утвердить.' % blok,
        'Ответственным руководителем массового взрыва %s назначить: %s.' % (blok, nu),
        'Ответственность за вывод людей за пределы опасной зоны взрывных работ '
        '(%s метров по проекту) возложить на: %s.' % (calc["zona_ludi"], nu),
        'Ответственность за вывод техники за пределы опасной зоны (%s метров по '
        'проекту) возложить на: %s.' % (calc["zona_obor"], nu),
        'Непосредственное руководство работой персонала при подготовке и '
        'проведении массового взрыва, учет и сохранность ВМ при заряжании '
        'скважин возложить на: %s.' % nu,
        'Взрывником назначить: Старший взрывник %s.' % vzr,
        'Директор прииска обязан обеспечить подготовку и проведение массового '
        'взрыва: автотранспортом, рабочим персоналом для забойки скважин и '
        'загрузки АС в СЗМ, а также для назначения на посты охраны опасной зоны.',
        'Ответственность за отключение электроэнергии возложить на: %s.' % nu,
        'Зарядку скважин произвести: %s с %s до %s (сырые скважины зарядить '
        'с использованием полиэтиленового рукава).' % (D, p["zar1"], p["zar2"]),
        'Взрыв произвести: %s с %s до %s.' % (D, p["vz1"], p["vz2"]),
        'Оповещение ответственного руководителя взрывных работ %s об окончании '
        'взрывных работ или изменениях даты и времени взрыва, с объяснением '
        'причин, возложить на: ответственного руководителя ВР %s.' % (pred, nu),
        'Оповещение диспетчерской службы Зонального центра и Якутского районного '
        'центра единой системы организации воздушного движения о сроках и времени '
        'проведения массового взрыва возложить на: Инженер ОТ и ПБ %s.' % injot,
        'Ответственность за проведение инструктажа персоналу, привлекаемому к '
        'работам по загрузке АС в СЗМ, забойке скважин и охране опасной зоны, '
        'возлагается на: %s.' % nu,
        'Ответственный руководитель ВР, %s, должен вести работы в соответствии '
        'с утвержденным распорядком проведения массового взрыва, требованиями '
        'ФНиП ПБ «Правила безопасности при взрывных работах» и инструкций.' % nu,
        'Оповещение диспетчера %s о времени проведения массового взрыва за 30 мин '
        'до начала МВ — %s.' % (pred, dispet),
        'Все назначенные настоящим приказом ответственные лица на период '
        'подготовки и производства взрывных работ подчиняются ответственному '
        'руководителю массового взрыва.',
        'Контроль за исполнением настоящего приказа оставляю за собой.',
    ]
    for i, it in enumerate(pitems, 1):
        AP("%d. %s" % (i, it))
    prikaz.append(Spacer(1, 14))
    prikaz.append(Paragraph("Главный инженер %s &nbsp;&nbsp;&nbsp; ________________ %s" % (pred, gling), pst))
    story.append(KeepTogether(prikaz))
    story.append(PageBreak())

    # ================= РАСПОРЯЖЕНИЕ О ПОСТАХ (лист 11) ===============
    P("Р А С П О Р Я Ж Е Н И Е", h1)
    P("о расстановке постов оцепления по охране опасной зоны", center)
    P("ВР на месторождении %s, блок %s" % (mest, blok), center)
    SP(6)
    P("Взрыв %s, время взрыва с %s до %s часов." % (D, p["vz1"], p["vz2"]))
    P("%s, в целях обеспечения безопасности производства взрыва, производит "
      "расстановку постов оцепления на границах опасной зоны %s м, в местах, "
      "указанных на плане горных работ, в следующем порядке:" % (nu, calc["zona_ludi"]))
    SP(2)
    for i in range(1, 6):
        P("Пост №%d — перекрывает подходы со стороны ________________________." % i, item)
        P("Ответственный ____________________ (подпись) ____________________", item)
    SP(4)
    P("ВСЕМ ПОСТОВЫМ!", bld)
    P("1. Бдительно наблюдать за местностью, самовольно не покидать пост "
      "и никого не пропускать в опасную зону.", item)
    P("2. Внимательно следить за подаваемыми сигналами:", item)
    P("ОДИН ДЛИННЫЙ звуковой сигнал — предупредительный;", center)
    P("ДВА ДЛИННЫХ звуковых сигнала — боевой;", center)
    P("ТРИ КОРОТКИХ звуковых сигнала — отбой.", center)
    SP(10)
    P("Ответственный за выставление постов опасной зоны: %s _____________ %s" % (nu, D))
    P("%s %s _____________ %s" % (nu, pred, D))
    story.append(PageBreak())

    # =================== РАСЧЁТ ПАРАМЕТРОВ (лист 12) =================
    P("Расчет параметров буровзрывных работ и потребности материалов", cbold)
    SP(4)
    P("Вместимость погонного метра скважины (Р) определяется по справочным "
      "таблицам исходя из диаметра скважины и плотности ВВ в заряде и "
      "составит: <b>%s</b> кг/пог.м." % fmt(calc["raskh_sel"]))
    P("Линия сопротивления по подошве уступа (ЛСПП) для вертикального заряда "
      "рассчитывается по формуле:")
    P("Wп = 53 · kт · dзар · √((ρвв·1000 − l) / ρпор), м", center)
    P("где: kт — коэффициент местных геологических условий, kт = %s; "
      "dзар — диаметр заряда, dзар = %s м; ρвв — плотность ВВ в заряде, "
      "ρвв = %s г/см³; l — коэффициент работоспособности ВВ (игданит) по "
      "отношению к %s, l = %s; ρпор — плотность породы, ρпор = %s кг/м³."
      % (fmt(p["ktresh"], 1), fmt(p["diam"], 3), fmt(p["plotvv"]),
         bname, fmt(p["krab"]), fmt0(p["plotpor"])))
    P("Wп = 53 · %s · %s · √((%s − %s) / %s) = <b>%s</b> м."
      % (fmt(p["ktresh"], 1), fmt(p["diam"], 3), fmt0(p["plotvv"] * 1000),
         fmt(p["krab"]), fmt0(p["plotpor"]), fmt(calc["lspp"], 1)), center)
    P("Расстояние между скважинами в ряду: a = Wп · m = %s · %s ≈ <b>%s</b> м; "
      "между рядами b ≈ <b>%s</b> м, где m — коэффициент сближения зарядов, "
      "m = %s." % (fmt(calc["lspp"], 1), fmt(calc["ksbl"]), fmt(p["a"], 1),
                   fmt(calc["b"], 1), fmt(calc["ksbl"])))
    P("Для инициирования зарядов скважин используются неэлектрические системы "
      "ИСКРА С/П, поэтому расстояние между рядами и скважинами принято равным.")
    P("Число рядов скважинных зарядов принято в соответствии с минимальной "
      "шириной рабочей площадки экскаватора. Количество рядов на полигоне %s "
      "составит <b>%s</b> шт." % (blok, p["ryad"]))
    P("Величина перебура: Lпер = 14 · dзар · √(ρвв · 0,8) = %s м; "
      "принимается равной <b>%s</b> м." % (fmt(calc["l_per_form"]), fmt(calc["sr_pereb"], 1)))
    story.append(PageBreak())

    # =================== РАСЧЁТ ПАРАМЕТРОВ (лист 13) =================
    P("Глубина скважины устанавливается по формуле:")
    P("Lс = H + Lпер = %s + %s = <b>%s</b> м."
      % (fmt(calc["sr_hust"]), fmt(calc["sr_pereb"]), fmt(calc["sr_glub"])), center)
    P("Коэффициент использования скважины (Кис):")
    P("Кис = H / Lс = %s / %s = <b>%s</b>."
      % (fmt(calc["sr_hust"]), fmt(calc["sr_glub"]), fmt(p["kis"])), center)
    P("Вес заряда в скважине (Qскв) принимается по фактическим данным каталога "
      "скважин:")
    P("Qскв = Qас/скв + Qдт/скв + Qбоевика = %s + %s + %s = <b>%s</b> кг."
      % (fmt(asn), fmt(dtn), fmt(p["boevik"]), fmt(calc["sr_massa"])), center)
    P("Длина заряда: Lзар = ΣLзар / nскв = <b>%s</b> м." % fmt(calc["sr_zar"]))
    P("Длина забойки: Lзаб = ΣLзаб / nскв = <b>%s</b> м." % fmt(calc["sr_zaboi"]))
    P("Площадь, приходящаяся на одну скважину: Sскв = a · b = %s · %s = "
      "<b>%s</b> кв.м." % (fmt(p["a"], 1), fmt(calc["b"], 1), fmt(calc["s_per_skv"])))
    P("Выход горной массы по маркшейдерским данным: Vгм/пог.м = Vмаркш / ΣLс = "
      "%s / %s = <b>%s</b> м³/пог.м." % (fmt0(calc["objem_massiva"]),
       fmt0(calc["sumD"]), fmt(calc["vyhod_gm"])))
    P("Объем взрываемой горной массы: Vгм = <b>%s</b> м³." % fmt(calc["objem_massiva"], 1))
    story.append(PageBreak())

    # =================== РАСЧЁТ ПАРАМЕТРОВ (лист 14) =================
    P("Суммарный расход ВВ (игданита) на взрыв составит:", bld)
    P("Qселитры = Nскв · Qскв = %d · %s = %s т;" % (n_disp, fmt(asn), fmt(calc["mas_as"] / 1000)))
    P("Qдт = Qселитры · %s = %s т;" % (fmt(p["doldt"], 3), fmt(calc["mas_dt_kg"] / 1000)))
    P("Mгранулита = %s + %s = %s т."
      % (fmt(calc["mas_as"] / 1000), fmt(calc["mas_dt_kg"] / 1000),
         fmt((calc["mas_as"] + calc["mas_dt_kg"]) / 1000)))
    P("Расход СИ: Искра-С — %s шт; Искра-П (67 мс) — %s шт; Искра-П (42 мс) — "
      "%s шт; Искра-В (Старт-600м) — %s шт." % (calc["iskra_s"], calc["iskra67"],
       calc["iskra42"], calc["iskraV"]))
    P("Расход %s на боевики: Mб · Nскв = %s · %d = %s кг."
      % (bname, fmt(p["boevik"]), n_disp, fmt(calc["mas_pt"])))
    SP(6)
    P("Основные параметры буровзрывных работ", cbold)
    op = [
        ("1", "Площадь полигона", "тыс.кв.м", fmt(p["ploshad"] / 1000, 3)),
        ("2", "Высота уступа", "м", fmt(calc["sr_hust"])),
        ("3", "Объем взрыва", "тыс.куб.м", fmt(calc["objem_massiva"] / 1000)),
        ("4", "Средняя глубина скважин", "м", fmt(avg_disp)),
        ("5", "в т.ч. перебур", "м", fmt(calc["sr_pereb"])),
        ("6", "Диаметр скважины", "мм", diam_mm),
        ("7", "Коэффициент использования скважины", "", fmt(p["kis"])),
        ("8", "Линия сопротивления по подошве (ЛСПП)", "м", fmt(calc["lspp"], 1)),
        ("9", "Расстояние между скважинами", "м", fmt(p["a"], 1)),
        ("10", "Расстояние между рядами скважин", "м", fmt(calc["b"], 1)),
        ("11", "Расположение скважин", "", p["raspol"]),
        ("12", "Количество скважин", "шт", str(n_disp)),
        ("13", "Объем бурения скважин", "пог.м", fmt0(drill_disp)),
        ("14", "Удельный расход Гранулита М", "кг/пог.м", fmt(calc["ud_vv"], 3)),
        ("15", "Длина заряда", "м", fmt(calc["sr_zar"])),
        ("16", "Длина забойки", "м", fmt(calc["sr_zaboi"])),
        ("17", "Величина заряда в скважине", "кг", fmt(calc["sr_massa"])),
        ("18", "Выход горной массы на 1 пог.м скважины", "куб.м", fmt(calc["vyhod_gm"])),
        ("19", "Расход Гранулит М", "т", fmt((calc["mas_as"] + calc["mas_dt_kg"]) / 1000, 3)),
        ("20", "Удельный расход АС на 1 куб.м ГМ", "кг/куб.м", fmt(calc["ud_as"], 3)),
        ("21", bname, "кг", fmt(calc["mas_pt"])),
        ("22", "Расход ИСКРА-С", "шт", str(calc["iskra_s"])),
        ("23", "Расход Искра-П (67 мс)", "шт", str(calc["iskra67"])),
        ("24", "Расход Искра-П (42 мс)", "шт", str(calc["iskra42"])),
        ("25", "Расход Искра-В (Старт-600м)", "шт", str(calc["iskraV"])),
        ("26", "Расход аммиачной селитры на взрыв", "т", fmt(calc["mas_as"] / 1000, 3)),
        ("27", "Удельный расход аммиачной селитры", "кг/м3", fmt(calc["ud_as"], 3)),
    ]
    tdata = [["№\nп/п", "Параметр", "Ед. изм.", "Средний\nпоказатель"]] + [list(r) for r in op]
    t = Table(tdata, colWidths=[28, 270, 80, 90])
    t.setStyle(TableStyle([
        ("FONT", (0, 0), (-1, -1), FONT, 8.5),
        ("FONT", (0, 0), (-1, 0), FONT_B, 8.5),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EEF3F9")),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (2, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(t)
    story.append(PageBreak())

    # ================= БЕЗОПАСНЫЕ РАССТОЯНИЯ (лист 15) ===============
    P("Расчет безопасных расстояний при взрывных работах", cbold)
    P("Расчет производится согласно главе XII ФНП «Правила безопасности при "
      "производстве, хранении и применении взрывчатых материалов промышленного "
      "назначения».")
    P("Определение зон, опасных по разлету отдельных кусков породы (грунта).", bld)
    P("Расстояние, безопасное для людей при взрывании скважинных зарядов, "
      "рассчитанных на дробящее действие, определяется по формуле:")
    P("rраз = 1250 · nз · √(4d / ((1+1)·b)), м", center)
    P("где: d — диаметр взрываемой скважины, d = %s м; b — расстояние между "
      "рядами, b = %s м; nз — коэффициент заполнения скважин ВВ."
      % (fmt(p["diam"], 3), fmt(calc["b"], 1)))
    P("nз = Lз / Lс = %s / %s = <b>%s</b>."
      % (fmt(calc["sr_zar"]), fmt(calc["sr_glub"]), fmt(calc["nz"], 3)))
    P("rраз = 1250 · %s · %s = <b>%s</b> м."
      % (fmt(calc["nz"], 3), fmt(calc["h9"]), fmt(calc["r_raz"], 1)), center)
    P("Коэффициент рельефа Кр = 0,5 · (1 + √(1 + 4H/rраз)) = <b>%s</b> "
      "(превышение H = %s м)." % (fmt(calc["kr"]), fmt(p["prev"], 1)))
    P("Rраз = rраз · Кр = %s · %s = <b>%s</b> м."
      % (fmt(calc["r_raz"], 1), fmt(calc["kr"]), fmt(calc["R_raz"], 1)), center)
    import math as _m
    P("Окончательное значение радиуса опасной зоны по разлету отдельных кусков "
      "породы принимается равным <b>%s</b> м." % max(1000, _m.ceil(calc["R_raz"] / 100) * 100))
    story.append(PageBreak())

    # ================= БЕЗОПАСНЫЕ РАССТОЯНИЯ (лист 16) ===============
    P("Определение расстояний, безопасных по действию ударной воздушной волны "
      "(УВВ) на здания и сооружения.", bld)
    P("Rв = 10 · Q^(1/3), где Q — максимальная масса заряда ВВ взрываемого блока, "
      "Q = %s кг." % fmt0(calc["obsh_massa"]))
    P("Rв = 10 · %s^(1/3) = <b>%s</b> м."
      % (fmt0(calc["obsh_massa"]), fmt(calc["r_uvv"], 1)), center)
    P("Принимаем величину безопасного расстояния по действию УВВ для зданий и "
      "сооружений равной <b>%s</b> м." % calc["r_uvv_zd"])
    P("Расчет сейсмически безопасных расстояний.", bld)
    P("Q = 2 · %s · 0,9 + 1 = %s кг." % (fmt(calc["sr_massa"], 1), fmt0(calc["q_seism"])))
    P("rс = 15,14 · Q^(1/3) = 15,14 · %s^(1/3) = <b>%s</b> м."
      % (fmt0(calc["q_seism"]), fmt(calc["r_seism"], 1)), center)
    P("Безопасное расстояние по сейсмическому действию принимаем равным <b>%s</b> м."
      % (_m.ceil(calc["r_seism"] / 10) * 10))
    P("Расчет расстояний, безопасных по высоте разлета отдельных кусков породы.", bld)
    P("rразл.в = 1,4 · rраз = 1,4 · %s = <b>%s</b> м."
      % (fmt(calc["r_raz"], 1), fmt(calc["r_raz_v"], 1)), center)
    P("Расстояние, безопасное по высоте разлета отдельных кусков породы, "
      "принимаем равным <b>%s</b> м." % max(500, _m.ceil(calc["r_raz_v"] / 100) * 100))
    story.append(PageBreak())

    # ================= БЕЗОПАСНЫЕ РАССТОЯНИЯ (лист 17) ===============
    P("Расчет расстояния, исключающего передачу детонации при хранении ВМ, "
      "доставленных к месту работ.", bld)
    P("Расчет произведен в соответствии с требованиями п. 845 главы XII ФНиП ПБ "
      "«Правила безопасности при взрывных работах». Рассматриваются ВВ "
      "(гранулит М) и средства инициирования; Кд = 0,8.")
    P("Окончательное значение расстояния, исключающего передачу детонации при "
      "хранении ВМ, доставленных к месту работ, принимаем равным <b>15</b> м.")
    P("По результатам расчетов радиус опасной зоны на участке производства "
      "взрывных работ принимаем:", bld)
    P("для людей — <b>%s</b> м;" % calc["zona_ludi"])
    P("для оборудования — <b>%s</b> м." % calc["zona_obor"])
    P("Принятые расстояния должны быть не менее расчетных по разлету кусков, УВВ, "
      "сейсмическому воздействию и высоте разлета; при наличии более жестких "
      "требований проекта принимается большее значение.")
    SP(18)
    P('Расчет составил: Начальник участка "%s" _____________ / %s /' % (mest, p["nachuch"]))
    SP(12)
    P("Проект массового взрыва со всеми графическими материалами хранится в "
      "делах взрывного участка (цеха) до полной отработки взорванного блока "
      "(ФНиП ПБ «Правила безопасности при взрывных работах»).", small)
    if passport_path:
        P("Приложение: паспорт буровзрывных работ — %s." % os.path.basename(passport_path), small)

    doc = SimpleDocTemplate(path, pagesize=A4,
                            leftMargin=18 * mm, rightMargin=14 * mm,
                            topMargin=14 * mm, bottomMargin=14 * mm)
    doc.build(story)
    return path


# ==================== HTML ====================
def build_html_context(calc, summary=None):
    """Готовит контекст для печатного HTML-шаблона проекта (bvr/project.html).

    Возвращает уже отформатированные строки/числа, чтобы шаблон оставался
    «глупым». Сводные count/метраж/средняя глубина берутся из summary
    (паспорт), остальное — из расчёта по зарядной карте.
    """
    import math as _m

    p = calc["params"]
    s = summary or {}
    n_disp, drill_disp, avg_disp = _summary_values(calc, summary)
    _dd, _mon, yy = date_parts(p)

    asn = calc["mas_as"] / calc["n"] if calc["n"] else 0
    dtn = calc["mas_dt_kg"] / calc["n"] if calc["n"] else 0

    # --- Зарядная карта: сетка с нумерацией рядов и скважин ---
    max_skv = max((w["skv"] for w in calc["wells"]), default=0)
    max_ryad = max((w["ryad"] for w in calc["wells"]), default=0)
    cell = {(w["ryad"], w["skv"]): w["d"] for w in calc["wells"]}
    grid_cols = list(range(1, max_skv + 1))
    grid_rows = []
    for r in range(1, max_ryad + 1):
        cells = []
        for c in grid_cols:
            v = cell.get((r, c))
            cells.append(fmt(v, 1) if v else "")
        grid_rows.append({"ryad": r, "cells": cells})

    # --- Таблица параметров по группам глубин ---
    depth_table = []
    for i, g in enumerate(depth_groups(calc), 1):
        mass = (g["sK"] + g["sL"]) / g["n"] + p["boevik"]
        depth_table.append({
            "i": i, "d": fmt(g["d"], 1), "n": g["n"],
            "a": fmt(p["a"], 1), "b": fmt(calc["b"], 1),
            "diam": round(p["diam"] * 1000), "mass": fmt(mass, 1),
            "zar": fmt(g["sN"] / g["n"], 1), "zab": fmt(g["sZ"] / g["n"], 1),
        })

    # --- Основные параметры (как в PDF, count/метраж/средняя — из summary) ---
    osn = [
        (1, "Площадь полигона", "тыс.кв.м", fmt(p["ploshad"] / 1000, 1)),
        (2, "Высота уступа", "м", fmt(calc["sr_hust"])),
        (3, "Объём взрыва", "тыс.куб.м", fmt(calc["objem_massiva"] / 1000)),
        (4, "Средняя глубина скважин", "м", fmt(avg_disp)),
        (5, "в т.ч. перебур", "м", fmt(calc["sr_pereb"])),
        (6, "Диаметр скважины", "мм", str(round(p["diam"] * 1000))),
        (7, "КИС", "", fmt(p["kis"])),
        (8, "ЛСПП", "м", fmt(calc["lspp"], 1)),
        (9, "Расстояние между скважинами", "м", fmt(p["a"], 1)),
        (10, "Расстояние между рядами", "м", fmt(calc["b"], 1)),
        (11, "Расположение скважин", "", p["raspol"]),
        (12, "Количество скважин", "шт", str(n_disp)),
        (13, "Объём бурения скважин", "пог.м", fmt0(drill_disp)),
        (14, "Удельный расход Гранулита М", "кг/пог.м", fmt(calc["ud_vv"], 3)),
        (15, "Длина заряда", "м", fmt(calc["sr_zar"])),
        (16, "Длина забойки", "м", fmt(calc["sr_zaboi"])),
        (17, "Величина заряда в скважине", "кг", fmt(calc["sr_massa"])),
        (18, "Выход горной массы на 1 пог.м", "куб.м", fmt(calc["vyhod_gm"])),
        (19, "Расход Гранулита М", "т",
         fmt((calc["mas_as"] + calc["mas_dt_kg"]) / 1000, 3)),
        (20, "Удельный расход АС на 1 куб.м ГМ", "кг/куб.м", fmt(calc["ud_as"], 3)),
        (21, _boevik_name(p), "кг", fmt(calc["mas_pt"])),
        (22, "Расход ИСКРА-С", "шт", str(calc["iskra_s"])),
        (23, "Расход Искра-П (67 мс)", "шт", str(calc["iskra67"])),
        (24, "Расход Искра-П (42 мс)", "шт", str(calc["iskra42"])),
        (25, "Расход Искра-В (Старт-600м)", "шт", str(calc["iskraV"])),
        (26, "Расход аммиачной селитры на взрыв", "т", fmt(calc["mas_as"] / 1000, 3)),
        (27, "Удельный расход аммиачной селитры", "кг/м3", fmt(calc["ud_as"], 3)),
    ]
    osn_params = [{"i": a, "name": b, "unit": c, "val": d} for (a, b, c, d) in osn]

    safety = {
        "nz": fmt(calc["nz"], 3), "h9": fmt(calc["h9"]),
        "r_raz": fmt(calc["r_raz"], 1), "kr": fmt(calc["kr"]),
        "R_raz": fmt(calc["R_raz"], 1),
        "r_raz_round": max(1000, _m.ceil(calc["R_raz"] / 100) * 100),
        "r_uvv": fmt(calc["r_uvv"], 1), "r_uvv_zd": calc["r_uvv_zd"],
        "q_seism": fmt0(calc["q_seism"]), "r_seism": fmt(calc["r_seism"], 1),
        "r_raz_v": fmt(calc["r_raz_v"], 1),
        "zona_ludi": calc["zona_ludi"], "zona_obor": calc["zona_obor"],
        "zoso": p["zoso"],
    }

    grid_mismatch = bool(s.get("from_passport")) and (
        n_disp != calc["n"]
        or abs(float(drill_disp) - float(calc["sumD"])) > 0.5)

    return {
        "p": p,
        "date_str": date_str(p),
        "year": yy,
        "summary": {
            "wells": n_disp,
            "meters": fmt(drill_disp, 1),
            "avg": fmt(avg_disp),
            "volume": fmt(calc["objem_massiva"] / 1000, 2),
            "mass": fmt(calc["obsh_massa"], 1),
            "zona_ludi": calc["zona_ludi"],
            "zona_obor": calc["zona_obor"],
            "from_passport": bool(s.get("from_passport")),
        },
        "grid_cols": grid_cols,
        "grid_rows": grid_rows,
        "grid_mismatch": grid_mismatch,
        "grid_wells": calc["n"],
        "grid_meters": fmt(calc["sumD"], 1),
        "depth_table": depth_table,
        "osn_params": osn_params,
        "safety": safety,
        "asn": fmt(asn), "dtn": fmt(dtn),
        "sr_massa": fmt(calc["sr_massa"]),
        "mas_pt": fmt(calc["mas_pt"], 1),
        "granulit_t": fmt((calc["mas_as"] + calc["mas_dt_kg"]) / 1000, 3),
        "lspp": fmt(calc["lspp"], 1),
        "ksbl": fmt(calc["ksbl"]),
        "raskh_sel": fmt(calc["raskh_sel"]),
        "vyhod_gm": fmt(calc["vyhod_gm"]),
        "ploshad_tm2": fmt(p["ploshad"] / 1000, 1),
        "boevik_name": _boevik_name(p),
        "participants": [{"fio": f, "prof": pr} for (f, pr) in _participants(p)],
    }
